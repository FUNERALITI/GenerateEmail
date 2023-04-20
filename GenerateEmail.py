import openpyxl as op
import random
from validate_email import validate_email
import time

# Подсчет времени работы программы
start = time.time()


# Конвертирование русских букв в английские
def convert_to_english(name):
    translation = {'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo', 'ж': 'zh', 'з': 'z', 'и': 'i',
                   'й': 'y',
                   'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
                   'ф': 'f',

                   'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch', 'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e',
                   'ю': 'yu',
                   'я': 'ya'}

    english_name = ''.join(translation.get(char, char) for char in name.lower())
    return english_name


# Генерация почтового ящика
def generate_mail(*ru_name):
    emails = []
    for name in ru_name:
        eng_name = convert_to_english(name)
        domains = ["@mail.ru", "@bk.ru", "@inbox.ru", "@gmail.com"]
        domain = random.choice(domains)
        postal_code = random.randint(1, 1000)
        email = f"{eng_name}{postal_code}{domain}"
        emails.append(email)
    return emails


# Проверка на существование
def validate(*lst):
    l_mail = []
    emails = generate_mail(*lst)
    for valid in emails:
        is_valid = validate_email(valid, verify=True)
        if is_valid:
            l_mail.append(valid)
    return l_mail


# Логика
def main():
    book = op.load_workbook(filename="WORKS.xlsx")
    sheet = book.active
    lst = []
    for row in range(2, sheet.max_row + 1):
        srt = sheet[row][1].value.split()
        s = "_".join(srt[0:2])
        lst.append(s)
    emails = validate(*lst)
    m_row = sheet.min_row + 1
    for i, email in enumerate(emails):
        sheet.cell(row=m_row + i, column=3).value = email
    book.save("WORKS.xlsx")


if __name__ == '__main__':
    main()


end = time.time() - start
print("Время работы программы: ", end)